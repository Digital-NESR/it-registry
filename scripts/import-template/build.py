import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.comments import Comment

here = os.path.dirname(os.path.abspath(__file__))
spec = json.load(open(os.path.join(here, "spec.json"), encoding="utf-8"))
cc   = json.load(open(os.path.join(here, "cc.json"), encoding="utf-8"))

countries = cc["countries"]; companies = cc["companies"]
compByCountry = cc["companiesByCountry"]; ccByCompany = cc["ccByCompany"]; ccDept = cc["ccDept"]

DATA_ROWS = 300
FIRST = 2
LAST = FIRST + DATA_ROWS - 1

wb = Workbook()

# ---------- Lists sheet (hidden) ----------
lists = wb.active; lists.title = "Lists"
col = 1
def write_list(header, values):
    global col
    L = get_column_letter(col)
    lists.cell(row=1, column=col, value=header).font = Font(bold=True)
    vals = list(values) if values else [""]
    for i, v in enumerate(vals):
        lists.cell(row=2 + i, column=col, value=v)
    rng = "Lists!$%s$2:$%s$%d" % (L, L, 1 + len(vals))
    col += 1
    return rng

def add_name(name, ref):
    wb.defined_names.add(DefinedName(name=name, attr_text=ref))

add_name("CountryList", write_list("Countries", countries))
for i, c in enumerate(countries, start=1):
    add_name("coco_%d" % i, write_list("CO:" + c[:20], compByCountry[c]))
add_name("CompanyList", write_list("Companies", companies))
for i, co in enumerate(companies, start=1):
    add_name("cc_%d" % i, write_list("CC:" + co[:20], ccByCompany[co]))
cc_codes = list(ccDept.keys())
add_name("CCcodes", write_list("CC codes", cc_codes))
add_name("CCdepts", write_list("CC depts", [ccDept[k] for k in cc_codes]))
for c in spec:
    if c["type"] == "dropdown":
        add_name("opt_" + c["key"], write_list(c["label"][:28], c["options"] or []))
lists.sheet_state = "hidden"

# ---------- Applications sheet ----------
ws = wb.create_sheet("Applications", 0)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
req_fill = PatternFill("solid", fgColor="C0392B")
opt_fill = PatternFill("solid", fgColor="2C3E50")
auto_fill = PatternFill("solid", fgColor="7F8C8D")
hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

letters = {c["key"]: get_column_letter(i + 1) for i, c in enumerate(spec)}

for i, c in enumerate(spec):
    L = get_column_letter(i + 1)
    cell = ws["%s1" % L]
    cell.value = c["label"] + (" *" if c["required"] else "")
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.fill = auto_fill if c["type"] == "cascade_dept" else (req_fill if c["required"] else opt_fill)
    cell.border = border
    note = (c.get("hint") or "") + ("\n(Required)" if c["required"] else "")
    if note.strip():
        cell.comment = Comment(note, "Template")
    ws.column_dimensions[L].width = max(12, min(30, len(c["label"]) + 4))

ws.row_dimensions[1].height = 34
ws.freeze_panes = "A2"

def dv_list(formula, rng):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=True)
    dv.error = "Pick a value from the dropdown list."; dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv); dv.add(rng)

for i, c in enumerate(spec):
    L = get_column_letter(i + 1)
    rng = "%s%d:%s%d" % (L, FIRST, L, LAST)
    t = c["type"]
    if t == "dropdown":
        dv_list("=opt_" + c["key"], rng)
    elif t == "cascade_country":
        dv_list("=CountryList", rng)
    elif t == "cascade_company":
        ctry = letters["country"]
        dv_list('=INDIRECT("coco_"&MATCH($%s%d,CountryList,0))' % (ctry, FIRST), rng)
    elif t == "cascade_cc":
        comp = letters["companyName"]
        dv_list('=INDIRECT("cc_"&MATCH($%s%d,CompanyList,0))' % (comp, FIRST), rng)
    elif t == "cascade_dept":
        ccl = letters["costCentre"]
        for r in range(FIRST, LAST + 1):
            ws["%s%d" % (L, r)] = '=IFERROR(INDEX(CCdepts,MATCH($%s%d,CCcodes,0)),"")' % (ccl, r)
    elif t == "date":
        dv = DataValidation(type="date", operator="between", formula1="DATE(1990,1,1)",
                            formula2="DATE(2100,12,31)", allow_blank=True, showErrorMessage=True)
        dv.error = "Enter a valid date."; dv.errorTitle = "Invalid date"
        ws.add_data_validation(dv); dv.add(rng)
        for r in range(FIRST, LAST + 1):
            ws["%s%d" % (L, r)].number_format = "yyyy-mm-dd"
    elif t == "number":
        dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0",
                            allow_blank=True, showErrorMessage=True)
        dv.error = "Enter a number (0 or greater)."; dv.errorTitle = "Invalid number"
        ws.add_data_validation(dv); dv.add(rng)

body_font = Font(name="Arial", size=10)
grey_fill = PatternFill("solid", fgColor="ECF0F1")
for r in range(FIRST, LAST + 1):
    for i in range(len(spec)):
        cl = ws.cell(row=r, column=i + 1)
        cl.font = body_font
        if spec[i]["type"] == "cascade_dept":
            cl.fill = grey_fill

# ---------- Example row ----------
country0 = "United Arab Emirates" if "United Arab Emirates" in countries else countries[0]
company0 = compByCountry[country0][0]
cc0 = ccByCompany[company0][0]
sourcing0 = (next((c for c in spec if c["key"] == "sourcing"), {}).get("options") or [""])[0]
ex = {"name": "Example: Contract Management System", "alias": "CMS",
      "country": country0, "companyName": company0, "costCentre": cc0, "sourcing": sourcing0}
example_font = Font(name="Arial", size=10, italic=True, color="7F8C8D")
for i, c in enumerate(spec):
    L = get_column_letter(i + 1); k = c["key"]; t = c["type"]
    if t == "cascade_dept":
        continue
    if k in ex:
        val = ex[k]
    elif t == "dropdown":
        val = (c["options"] or [""])[0]
    elif t == "date":
        val = "2025-01-15"
    elif t == "number":
        val = 0
    else:
        val = ""
    if val != "" or k in ex:
        cell = ws["%s%d" % (L, FIRST)]; cell.value = val; cell.font = example_font
ws["A%d" % FIRST].comment = Comment("EXAMPLE ROW - overwrite or delete before importing.", "Template")

# ---------- Instructions sheet ----------
info = wb.create_sheet("Instructions", 0)
info.column_dimensions["A"].width = 110
title_f = Font(bold=True, size=14, name="Arial", color="2C3E50")
h_f = Font(bold=True, size=11, name="Arial", color="2C3E50")
b_f = Font(size=10, name="Arial")
lines = [
    ("NESR IT Application Registry - Import Template", title_f),
    ("", b_f),
    ("How to use this template", h_f),
    ("1. Fill in one application per row on the 'Applications' sheet, starting at row 3 (row 2 is an example - overwrite or delete it).", b_f),
    ("2. Cells with a dropdown arrow only accept values from the list. Values not in the list are rejected.", b_f),
    ("3. Column headers marked with * are required.", b_f),
    ("", b_f),
    ("Cascading location fields (fill left to right)", h_f),
    ("- Country: choose first.", b_f),
    ("- Company Name: list is filtered by the Country you picked.", b_f),
    ("- Cost Centre: list is filtered by the Company you picked.", b_f),
    ("- Department: filled in automatically from the Cost Centre (grey column - do not type here).", b_f),
    ("", b_f),
    ("Field types", h_f),
    ("- Dropdown fields: pick from the list.", b_f),
    ("- Date fields: enter as YYYY-MM-DD (e.g. 2025-01-15).", b_f),
    ("- Number/cost fields: numbers only (0 or greater). Leave blank or 0 if unknown.", b_f),
    ("- Yes/No fields: pick Yes or No.", b_f),
    ("", b_f),
    ("Not included in this template (add inside the app after import)", h_f),
    ("- File attachments (data-flow doc, certification certificates, general documents).", b_f),
    ("- Provider / partner contact lists and AI model lists.", b_f),
    ("- Certifications and upstream/downstream application links.", b_f),
    ("- Total Cost of Ownership is calculated automatically (License + Maintenance cost).", b_f),
    ("", b_f),
    ("Header colour key:  RED = required   |   DARK BLUE = optional   |   GREY = auto-filled", h_f),
]
for i, (txt, f) in enumerate(lines, start=1):
    cinfo = info.cell(row=i, column=1, value=txt); cinfo.font = f
    cinfo.alignment = Alignment(wrap_text=True, vertical="top")
info.sheet_view.showGridLines = False

wb._sheets.sort(key=lambda s: {"Instructions": 0, "Applications": 1, "Lists": 2}.get(s.title, 3))

out = os.path.join(here, "..", "..", "NESR_IT_Application_Import_Template.xlsx")
wb.save(out)
print("saved:", out)
print("columns:", len(spec), "| data rows:", DATA_ROWS)
